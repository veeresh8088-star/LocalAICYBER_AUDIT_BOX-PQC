# -*- coding: utf-8 -*-
"""
Excel Scoping Parser
====================
Dynamically parses any ISO 27001 audit scoping checklist from an Excel file.

Features:
- Auto-detects 1 to N file columns from the header row (keywords: file, evidence,
  policy, document, attachment, doc, upload)
- Resolves ISO control IDs using 3 fallback steps:
    1. Direct control ID match (e.g. "5.15", "ISO 8.17")
    2. Control name keyword match against USE_CASES
    3. Embedding cosine-similarity against USE_CASES descriptions (no LLM needed)
- Returns a list of dicts: {question, files, control_id, control_label,
  expected_evidence, prompt_hint, severity}
- Works for 8 rows, 50 rows, 100+ rows
"""

import re
import os
from typing import List, Dict, Optional

# ── Keywords that identify file-reference columns in the Excel header ─────────
# IMPORTANT: 'type' columns (File type, Document type) must be EXCLUDED because
# they contain extensions like 'PNG', 'JPG', 'PDF' which are not filenames.
_FILE_COL_KEYWORDS = {
    "file", "evidence", "policy", "document", "attachment", "doc", "upload",
    "exhibit", "reference", "source", "artifact"
}
# Words that, when present, indicate a TYPE/FORMAT column (not a filename column)
_FILE_TYPE_EXCLUSION_KEYWORDS = {"type", "format", "extension", "ext", "kind"}

# ── Direct keyword-to-control mapping for common audit questions ──────────────
# These cover the most frequent audit check phrasings so they resolve instantly
# without needing embedding similarity (which requires Ollama to be running).
_DIRECT_KEYWORD_CONTROL_MAP = [
    # Keywords in question text               -> control use_case string
    (["ntp", "clock", "synchroni", "time server"], "8.17 Clock Synchronization"),
    (["fraud analytics"], "5.1 Policies for Information Security"),
    (["multifactor", "mfa", "multi-factor", "2fa", "two-factor"], "8.5 Secure Authentication"),
    (["pam", "privileged access", "pim", "idam"], "8.2 Privileged Access Rights"),
    (["access control policy", "access control"], "5.15 Access Control"),
    (["authentication", "how is the auth"], "8.5 Secure Authentication"),
    (["asset management policy", "asset management", "asset inventory"], "5.9 Inventory of Information and Other Associated Assets"),
    (["incident management policy", "incident response", "incident plan", "irp"], "5.24 Information Security Incident Management Planning and Preparation"),
    (["business continuity plan", "bcp", "disaster recovery", "dr plan"], "5.29 Information Security During Disruption"),
    (["vulnerability", "patch", "scan"], "8.8 Management of Technical Vulnerabilities"),
    (["hr security policy", "people security policy", "screening"], "6.1 Screening"),
    (["physical security policy", "physical security perimeter", "physical perimeter", "perimeter security"], "7.1 Physical Security Perimeters"),
    (["security monitoring", "log monitoring", "siem", "monitoring activities", "monitoring"], "8.16 Monitoring Activities"),
    (["cpu", "memory", "disk", "utilization", "capacity", "cloudwatch"], "8.6 Capacity Management"),
    (["log archival", "log archived", "archiv", "records retention", "records management"], "5.33 Protection of Records"),
    (["backup", "recovery", "restore"], "8.13 Information Backup"),
    (["incident", "response", "breach"], "5.24 Information Security Incident Management Planning and Preparation"),
    # Matching is word-boundary anchored, so each surface form has to be listed:
    # "encryption" cannot match "encrypted", and the control's own name --
    # "Use of Cryptography" -- was itself unmatchable because neither
    # "cryptography" nor "cryptographic" appeared here. A checklist row reading
    # "Cryptographic Controls" therefore resolved to nothing under ISO, and
    # before framework confinement it reached SOC 2 CC3.4 instead.
    (["encryption", "encrypted", "encrypt", "cryptography", "cryptographic",
      "cryptographic controls", "key management", "key rotation", "kms",
      "tls", "ssl", "cipher", "aes", "data at rest", "data in transit"],
     "8.24 Use of Cryptography"),
    (["password", "credential", "secret"], "5.17 Authentication Information"),
    (["firewall", "network security", "network traffic", "firewall policy"], "8.20 Network Security"),
    (["gdpr", "pii", "personal data", "privacy"], "5.34 Privacy and Protection of Personally Identifiable Information (Pii)"),

    # ── Clause 5: Organizational Controls (all previously missing) ────────────
    (["information security policy", "isms policy", "security policy document", "iprotect"], "5.1 Policies for Information Security"),
    (["roles and responsibilities", "isms roles", "security roles", "responsibility assignment"], "5.2 Information Security Roles and Responsibilities"),
    (["segregation of duties", "separation of duties", "dual control", "four eyes principle"], "5.3 Segregation of Duties"),
    (["management commitment", "management responsibilities", "senior management", "executive sponsor"], "5.4 Management Responsibilities"),
    (["contact with authorities", "law enforcement", "regulatory body", "government contact", "police contact"], "5.5 Contact with Authorities"),
    (["special interest group", "isac", "industry group", "security forum", "peer group"], "5.6 Contact with Special Interest Groups"),
    (["threat intelligence", "threat feed", "cti", "threat data", "ioc", "indicators of compromise"], "5.7 Threat Intelligence"),
    (["project management", "project security", "sdlc governance", "project risk", "security in projects"], "5.8 Information Security in Project Management"),
    (["acceptable use", "aup", "usage policy", "acceptable use policy", "permitted use"], "5.10 Acceptable Use of Information and Other Associated Assets"),
    (["return of assets", "asset return", "exit assets", "equipment return", "device return on exit"], "5.11 Return of Assets"),
    (["data classification", "information classification", "classification scheme", "sensitivity label", "classification policy"], "5.12 Classification of Information"),
    (["data labelling", "information labelling", "label policy", "classification marking", "document marking"], "5.13 Labelling of Information"),
    (["information transfer", "data transfer", "file transfer", "email security", "secure transfer", "sftp protocol", "secure ftp"], "5.14 Information Transfer"),
    (["identity management", "user provisioning", "joiner leaver", "account lifecycle", "iam policy", "user lifecycle"], "5.16 Identity Management"),
    (["access rights", "user permissions", "permission management", "access provisioning", "least privilege access"], "5.18 Access Rights"),
    (["supplier relationship", "vendor relationship", "third party security", "supplier policy", "vendor agreement"], "5.19 Information Security in Supplier Relationships"),
    (["supplier agreement", "vendor contract", "third party contract", "outsourcing agreement", "service contract"], "5.20 Addressing Information Security Within Supplier Agreements"),
    (["ict supply chain", "supply chain security", "hardware supply", "software supply", "supply chain risk"], "5.21 Managing Information Security in The Ict Supply Chain"),
    (["supplier monitoring", "vendor monitoring", "third party review", "supplier audit", "vendor performance"], "5.22 Monitoring, Review and Change Management of Supplier Services"),
    (["cloud service", "cloud security", "saas", "iaas", "paas", "cloud provider", "aws", "azure", "gcp"], "5.23 Information Security for Use of Cloud Services"),
    (["incident assessment", "security event assessment", "incident triage", "event decision", "incident classification"], "5.25 Assessment and Decision on Information Security Events"),
    (["incident response plan", "incident handling", "containment", "eradication", "recovery response"], "5.26 Response to Information Security Incidents"),
    (["lessons learned", "post incident review", "incident review", "learning from incidents", "incident debrief"], "5.27 Learning from Information Security Incidents"),
    (["evidence collection", "forensic", "digital forensic", "chain of custody", "forensic evidence"], "5.28 Collection of Evidence"),
    (["ict continuity", "ict readiness", "it continuity", "system continuity", "technology continuity"], "5.30 Ict Readiness for Business Continuity"),
    (["legal requirement", "statutory requirement", "regulatory requirement", "contractual requirement", "compliance obligation"], "5.31 Legal, Statutory, Regulatory and Contractual Requirements"),
    (["intellectual property", "copyright", "ip rights", "software license", "license management"], "5.32 Intellectual Property Rights"),
    (["independent review", "internal audit", "isms review", "third party audit", "external audit"], "5.35 Independent Review of Information Security"),
    (["compliance check", "policy compliance", "compliance with policy", "compliance review", "standards compliance"], "5.36 Compliance with Policies and Standards for Information Security"),
    (["operating procedures", "sop", "documented procedures", "standard operating procedure", "work instructions"], "5.37 Documented Operating Procedures"),

    # ── Clause 6: People Controls (all previously missing) ────────────────────
    (["employment terms", "terms and conditions", "employment contract", "job agreement", "offer letter"], "6.2 Terms and Conditions of Employment"),
    (["security awareness", "awareness training", "security training", "e-learning", "phishing awareness", "staff training"], "6.3 Information Security Awareness, Education and Training"),
    (["disciplinary", "disciplinary process", "misconduct", "policy violation", "disciplinary action"], "6.4 Disciplinary Process"),
    (["termination", "offboarding", "exit process", "leaver", "resignation", "dismissal", "account deactivation on exit"], "6.5 Responsibilities after Termination or Change of Employment"),
    (["nda", "non-disclosure", "confidentiality agreement", "non disclosure agreement", "confidentiality clause"], "6.6 Confidentiality or Non-disclosure Agreements"),
    (["remote work", "remote working", "work from home", "wfh", "telework", "home working", "vpn policy"], "6.7 Remote Working"),
    (["security event reporting", "report incident", "event reporting", "staff reporting", "how to report"], "6.8 Information Security Event Reporting"),

    # ── Clause 7: Physical Controls (all previously missing) ──────────────────
    (["physical entry", "entry control", "door access", "turnstile", "access control door", "entry point"], "7.2 Physical Entry"),
    (["secure office", "secure room", "secure facility", "server room", "data center access", "office security"], "7.3 Securing Offices, Rooms and Facilities"),
    (["physical monitoring", "cctv", "camera surveillance", "security camera", "video surveillance"], "7.4 Physical Security Monitoring"),
    (["environmental threat", "flood protection", "fire protection", "power protection", "environmental control", "uninterruptible power"], "7.5 Protecting against Physical and Environmental Threats"),
    (["secure area", "working in secure area", "restricted area policy", "secure zone", "clean area"], "7.6 Working in Secure Areas"),
    (["clean desk policy", "clear screen policy", "unattended workstation", "clear desk"], "7.7 Clear Desk and Clear Screen"),
    (["equipment siting", "equipment placement", "server placement", "equipment protection", "rack security"], "7.8 Equipment Siting and Protection"),
    (["off-premises", "assets off-premises", "remote equipment", "equipment offsite", "off site device"], "7.9 Security of Assets Off-premises"),
    (["storage media", "removable media", "usb", "hard drive disposal", "media handling", "removable storage"], "7.10 Storage Media"),
    (["supporting utilities", "power supply failure", "backup generator", "utility failure", "electricity supply", "uninterruptible power supply"], "7.11 Supporting Utilities"),
    (["cabling", "cable security", "network cabling", "structured cabling", "cable management"], "7.12 Cabling Security"),
    (["equipment maintenance", "maintenance schedule", "server maintenance", "hardware maintenance", "preventive maintenance"], "7.13 Equipment Maintenance"),
    (["secure disposal", "equipment disposal", "data destruction", "disk wipe", "degauss", "decommission"], "7.14 Secure Disposal or Re-use of Equipment"),

    # ── Clause 8: Technological Controls (all previously missing) ─────────────
    (["endpoint", "user endpoint", "laptop policy", "mobile device", "mdm", "byod", "endpoint security"], "8.1 User Endpoint Devices"),
    (["information access restriction", "need to know", "access restriction", "data access control"], "8.3 Information Access Restriction"),
    (["source code", "code repository", "git access", "repository access", "source code access", "github", "gitlab"], "8.4 Access to Source Code"),
    (["anti-malware", "edr", "malware protection", "endpoint protection", "av policy"], "8.7 Protection against Malware"),
    (["configuration management", "baseline configuration", "hardening", "cmdb", "config baseline", "secure configuration"], "8.9 Configuration Management"),
    (["data deletion", "information deletion", "secure erase", "data wiping", "data removal", "right to erasure"], "8.10 Information Deletion"),
    (["data masking", "anonymization", "pseudonymization", "masking policy", "data anonymisation"], "8.11 Data Masking"),
    (["dlp", "data leakage prevention", "data loss prevention", "data exfiltration", "information leakage"], "8.12 Data Leakage Prevention"),
    (["system redundancy", "failover mechanism", "high availability setup", "active active", "active passive", "redundant system"], "8.14 Redundancy of Information Processing Facilities"),
    (["log management", "syslog", "audit logs", "event logs", "logging policy", "log collection", "log retention"], "8.15 Logging"),
    (["privileged utility", "admin tools", "utility programs", "system utilities", "privileged software"], "8.18 Use of Privileged Utility Programs"),
    (["software installation", "approved software", "application whitelist", "install policy", "software approval"], "8.19 Installation of Software on Operational Systems"),
    (["network services", "service security", "api gateway", "network service policy", "managed network service"], "8.21 Security of Network Services"),
    (["network segregation", "vlan", "network segmentation", "dmz", "network zone", "micro segmentation"], "8.22 Segregation of Networks"),
    (["web filtering", "url filtering", "proxy", "content filter", "internet filtering", "web proxy"], "8.23 Web Filtering"),
    (["secure development", "sdlc", "secure development lifecycle", "secure development policy"], "8.25 Secure Development Life Cycle"),
    (["application security", "app security requirements", "security requirements", "security in design"], "8.26 Application Security Requirements"),
    (["secure architecture", "security architecture", "engineering principles", "security by design", "architecture review"], "8.27 Secure System Architecture and Engineering Principles"),
    (["secure coding", "code review", "coding standard", "owasp", "sast", "dast", "static analysis"], "8.28 Secure Coding"),
    (["security testing", "acceptance testing", "uat security", "pre-production testing", "security test"], "8.29 Security Testing in Development and Acceptance"),
    (["outsourced development", "third party development", "vendor development", "offshore development"], "8.30 Outsourced Development"),
    (["separation of environments", "dev test prod", "environment separation", "non-production", "staging environment"], "8.31 Separation of Development, Testing and Production Environments"),
    (["change management", "change control", "change request", "change advisory board", "change approval process"], "8.32 Change Management"),
    (["test information", "test data", "test data management", "production data in test", "sanitised test data"], "8.33 Test Information"),
    (["audit testing", "audit protection", "system during audit", "audit tools", "audit environment"], "8.34 Protection of Information Systems During Audit Testing"),
]


# Short tokens in _DIRECT_KEYWORD_CONTROL_MAP that are complete acronyms, not
# stems -- confirmed to bare-substring-collide with unrelated words: "pam" in
# "spam", "nda" in "agenda", "irp" in "airport", "edr" in "bedroom", "cti" in
# "practice", "aws" in "jaws". These need \b word-boundary matching. Other
# short-looking tokens in that map (e.g. "synchroni", "archiv") are
# deliberate stems meant to bare-substring-match "synchronization"/
# "synchronized"/"archived"/"archival" etc. -- \b would silently break that
# stemming, so only this specific confirmed-bad set gets the stricter check.
_WHOLE_WORD_ONLY_MAP_KEYWORDS = {"pam", "nda", "irp", "edr", "cti", "aws"}


def _resolve_control_by_direct_map(text: str, use_cases: List[Dict]) -> Optional[Dict]:
    """Fast O(n) keyword-map lookup before falling back to USE_CASES search.
    Ranks matches by keyword length so specific multi-word phrases (e.g. 'supplier monitoring')
    take precedence over shorter substring matches (e.g. 'monitoring')."""
    text_norm = _normalize(text)
    text_lower = str(text or "").lower()
    matches = []

    for keywords, target_use_case in _DIRECT_KEYWORD_CONTROL_MAP:
        for kw in keywords:
            kw_norm = _normalize(kw)
            if not kw_norm:
                continue
            if kw in _WHOLE_WORD_ONLY_MAP_KEYWORDS:
                hit = _kw_match(kw_norm, text_norm) or _kw_match(kw.lower(), text_lower)
            else:
                hit = kw_norm in text_norm or kw.lower() in text_lower
            if hit:
                matches.append((len(kw_norm), target_use_case))

    if matches:
        # Longest match wins (e.g. 'supplier monitoring' len 19 > 'monitoring' len 10)
        matches.sort(key=lambda x: x[0], reverse=True)
        best_target = matches[0][1]
        target_norm = _normalize(best_target)
        for uc in use_cases:
            if _normalize(uc.get("use_case", "")) == target_norm:
                return uc

    return None



# ── Control short-IDs across every framework in USE_CASES.
#
# This previously matched only "N.N" (ISO) and "VAPT-N", so Step 1 of
# _resolve_control() -- the exact control-ID lookup -- was dead for five of the
# eight frameworks. Their IDs simply never matched, and resolution silently fell
# through to fuzzy name matching. Two consequences, both confirmed by running the
# resolver over all 217 controls:
#
#   * An Excel checklist carrying only a Control-ID column resolved 1/15 for DPDP,
#     1/12 for PQC, 1/23 for XBOM, 1/4 for BCMS and 9/33 for SOC 2.
#   * With ID matching unavailable, the name matcher crossed framework boundaries:
#     NIST "PR.AT" resolved to ISO 6.3 and "RS.MA" to ISO 5.24 -- auditing a NIST
#     control against a different framework's text and expected evidence.
#
# ID matching short-circuits before the name matcher, so covering every real shape
# fixes both. The alternation is ordered longest-prefix-first: the alpha-dash
# families must be tried before the bare numeric pattern, or "XBOM-7" would match
# nothing while a stray "6.1" inside free text won as an ISO clause.
#
# Shapes present in USE_CASES: N.N (ISO, 93) | CCN.N (SOC 2, 33) | XBOM-N (23)
# | VAPT-N (15) | DPDP-N (15) | PQC-N (12) | BCMS-N (4) | XX.YY (NIST CSF, 22).
_CONTROL_ID_RE = re.compile(
    r'\b('
    r'(?:VAPT|DPDP|PQC|XBOM|BCMS)\s*-?\s*\d{1,3}'   # VAPT-1, DPDP-15, XBOM-23, PQC-5, BCMS-2
    r'|CC\d(?:\.\d{1,2})?'                          # SOC 2: CC6, CC6.1
    r'|(?:GV|ID|PR|DE|RS|RC)\.[A-Z]{2}'             # NIST CSF: GV.OC, PR.AT, RS.MA
    r'|\d{1,2}\.\d{1,2}(?:\.\d{1,2})?'              # ISO 27001: 5.1, 8.17, 8.17.1
    r')\b',
    re.IGNORECASE
)

# Frameworks whose IDs are written "<PREFIX>-<number>". Normalised to a single
# canonical dash so "dpdp 3", "DPDP3" and "DPDP - 3" all reach "DPDP-3".
_DASHED_ID_PREFIXES = ("VAPT", "DPDP", "PQC", "XBOM", "BCMS")


def _normalize(text: str) -> str:
    """Lowercase and strip punctuation for fuzzy matching.

    Collapses runs of whitespace down to a single space -- without this,
    "Hiring & Termination" (ampersand + surrounding spaces -> 3 consecutive
    spaces) and "Hiring__Termination" (double underscore, e.g. from the
    upload endpoint's filename sanitizer -> 2 consecutive spaces) normalize
    to strings that differ only in whitespace run-length, which breaks both
    exact and substring matching in _match_filenames even though they're
    clearly the same reference. Confirmed against a real Excel citation vs.
    its actual sanitized uploaded filename.
    """
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9\s]', ' ', str(text or "").lower())).strip()


def _kw_match(kw: str, text: str) -> bool:
    """Word-boundary keyword match. A bare "kw in text" substring check let
    short keywords match inside unrelated words -- confirmed: "id" inside
    "Evidence Provided" (misdetected as the Control ID column), "ext" inside
    "Extra Attachments" (wrongly excluded as a type/format column), "pam"
    inside "spam", "nda" inside "agenda", "irp" inside "airport", "edr"
    inside "bedroom", "cti" inside "practice" (all in
    _DIRECT_KEYWORD_CONTROL_MAP, which scans free-text audit question cells,
    not just headers). Both `text` and `kw` are expected already
    punctuation-normalized (via _normalize) so "_"/"-" don't break \b."""
    if not kw:
        return False
    return re.search(r'\b' + re.escape(kw) + r'\b', text) is not None


def _load_use_cases() -> List[Dict]:
    """Load USE_CASES from controls_data safely."""
    try:
        from src.core.controls_data import USE_CASES
        return USE_CASES
    except ImportError:
        return []


def _resolve_control_by_id(text: str, use_cases: List[Dict]) -> Optional[Dict]:
    """Step 1: Extract control ID directly from text (e.g., '8.16', '5.15', 'VAPT-1', 'ISO 8.17')."""
    if not text:
        return None
    matches = _CONTROL_ID_RE.findall(str(text))
    for m in matches:
        # Canonicalise "<PREFIX> <n>" / "<PREFIX>-<n>" / "<PREFIX><n>" to "PREFIX-n".
        # This handled VAPT alone before; the other dashed families fell through
        # unnormalised and only matched when the sheet already used the exact form.
        m_norm = m.strip().upper()
        for prefix in _DASHED_ID_PREFIXES:
            m_norm = re.sub(
                rf'^{prefix}\s*-?\s*(\d+)$', rf'{prefix}-\1', m_norm, flags=re.IGNORECASE
            )
        for uc in use_cases:
            uc_id = str(uc.get("use_case", "")).split(" ")[0].upper()
            if uc_id == m_norm or uc_id == m.upper():
                return uc
    return None




def _resolve_control_by_name(text: str, use_cases: List[Dict]) -> Optional[Dict]:
    """Step 2: Match control by name keywords from USE_CASES labels."""
    norm = _normalize(text)
    if not norm:
        return None

    words = [w for w in norm.split() if len(w) > 2]  # Filter tiny stop words
    if not words:
        return None

    best_uc = None
    best_score = 0

    for uc in use_cases:
        label = _normalize(uc.get("label", "") + " " + uc.get("use_case", ""))
        l_words = set(label.split())
        overlap = len(set(words) & l_words)
        if overlap > best_score:
            best_score = overlap
            best_uc = uc

    # If query is short (1-2 meaningful words like 'Capacity' or 'Screening'), require at least 1 match.
    # Otherwise require at least 2 matches.
    min_required = 1 if len(words) <= 2 else 2
    return best_uc if best_score >= min_required else None


def _resolve_control_by_embedding(text: str, use_cases: List[Dict]) -> Optional[Dict]:
    """Step 3: Cosine similarity against USE_CASES descriptions using cached embeddings."""
    try:
        import numpy as np
        import requests

        # Use Ollama embedding model
        ollama_url = os.environ.get("OLLAMA_HOST", "http://127.0.0.1:11434")
        if not ollama_url.startswith("http"):
            ollama_url = f"http://{ollama_url}"

        embed_model = os.environ.get("EMBED_MODEL", "nomic-embed-text")

        def get_embedding(t: str):
            try:
                resp = requests.post(
                    f"{ollama_url}/api/embeddings",
                    json={"model": embed_model, "prompt": t},
                    timeout=60

                )
                if resp.status_code == 200:
                    return np.array(resp.json().get("embedding", []), dtype=np.float32)
            except Exception:
                pass
            return None

        q_vec = get_embedding(text)
        if q_vec is None or len(q_vec) == 0:
            return None

        best_uc = None
        best_sim = -1.0

        for uc in use_cases:
            desc = f"{uc.get('use_case', '')} {uc.get('label', '')} {uc.get('prompt_hint', '')}"
            uc_vec = get_embedding(desc)
            if uc_vec is None or len(uc_vec) == 0:
                continue
            denom = (np.linalg.norm(q_vec) * np.linalg.norm(uc_vec))
            if denom == 0:
                continue
            sim = float(np.dot(q_vec, uc_vec) / denom)
            if sim > best_sim:
                best_sim = sim
                best_uc = uc

        return best_uc if best_sim > 0.5 else None

    except Exception as e:
        print(f"[EXCEL SCOPING] Embedding resolution failed: {e}", flush=True)
        return None
_STANDARD_ALIASES = {
    "ISO 27001": ("iso", "27001", "isms"),
    "SOC2":      ("soc2", "soc 2"),
    "XBOM":      ("xbom", "x-bom", "sbom"),
    "NIST CSF 2.0": ("nist", "csf"),
    "VAPT":      ("vapt", "pentest", "penetration"),
    "DPDP":      ("dpdp", "gdpr", "privacy"),
    "PQC":       ("pqc", "post-quantum", "quantum"),
    "BCMS":      ("bcms", "continuity", "22301"),
}


def filter_use_cases_by_framework(use_cases: List[Dict], framework: str) -> List[Dict]:
    """Narrows the candidate controls to the framework the audit is actually running.

    Without this, a checklist row carrying no control ID is matched by name or
    question against all 217 controls from all eight frameworks. The ID matcher
    was fixed to respect framework boundaries, but the name and question matchers
    below it were not -- so an ISO checklist row reading "Cryptographic Controls"
    resolved to SOC 2 "CC3.4 Identification of Changes Impacting Controls"
    instead of ISO "8.24 Use of Cryptography", and the auditor's crypto policy
    and evidence were then judged against a change-management requirement.

    An unrecognised or empty framework returns the list unchanged, so callers
    that do not know the framework behave exactly as before.
    """
    if not framework or not use_cases:
        return use_cases
    fw = str(framework).strip().lower()
    target = None
    for std, aliases in _STANDARD_ALIASES.items():
        if any(a in fw for a in aliases):
            target = std
            break
    if not target:
        return use_cases
    narrowed = [u for u in use_cases if str(u.get("standard", "")).strip() == target]
    # Never hand back an empty candidate set -- a filter that matches nothing
    # would make every row unresolvable, which is worse than cross-framework
    # matching. Fall back to the full list and let the matchers try.
    return narrowed or use_cases


def _resolve_control(
    id_text: str = "",
    name_text: str = "",
    q_text: str = "",
    use_cases: List[Dict] = None
) -> Dict:
    """
    Resolve ISO control from id/name/question text using 4 priority fallback steps:
    1. Control ID match
    2. Control Name match
    3. Question / Keyword match
    4. Cosine similarity embedding fallback
    """
    use_cases = use_cases or []
    id_text = str(id_text or "").strip()
    name_text = str(name_text or "").strip()
    q_text = str(q_text or "").strip()
    resolution_text = " ".join(t for t in [id_text, name_text, q_text] if t)

    # Priority 1: Control ID match
    uc = _resolve_control_by_id(id_text, use_cases) if id_text else None
    if not uc and resolution_text:
        uc = _resolve_control_by_id(resolution_text, use_cases)

    # Priority 2: Control Name match
    if not uc and name_text:
        uc = _resolve_control_by_name(name_text, use_cases)

    # Priority 3: Question / Keyword match
    if not uc and q_text:
        uc = _resolve_control_by_direct_map(q_text, use_cases)
    if not uc and resolution_text:
        uc = _resolve_control_by_direct_map(resolution_text, use_cases) or _resolve_control_by_name(resolution_text, use_cases)

    # Priority 4: Embedding fallback
    if not uc and resolution_text:
        uc = _resolve_control_by_embedding(resolution_text, use_cases)

    if uc:
        use_case_str = str(uc.get("use_case", ""))
        parts = use_case_str.split(" ", 1)
        ctrl_id = parts[0] if parts else "UNKNOWN"
        ctrl_name = parts[1] if len(parts) > 1 else use_case_str
        return {
            "control_id":        ctrl_id,
            "control_name":      ctrl_name,
            "control_label":     use_case_str,
            "expected_evidence": str(uc.get("expected", "")),
            "prompt_hint":       str(uc.get("prompt_hint", "")),
            "severity":          str(uc.get("severity", "MEDIUM")),
            "resolved":          True,
        }

    # Fallback: custom / unknown control
    fallback_name = name_text or id_text or q_text or "Unknown Control"
    fallback_id = id_text or "UNKNOWN"
    return {
        "control_id":        fallback_id,
        "control_name":      fallback_name,
        "control_label":     f"{fallback_id} {fallback_name}".strip() if fallback_id != "UNKNOWN" else fallback_name,
        "expected_evidence": "",
        "prompt_hint":       f"Evaluate the provided document against: {fallback_name}",
        "severity":          "MEDIUM",
        "resolved":          False,
    }


def _detect_file_columns(header_row: list) -> List[int]:
    """
    Auto-detect which column indices contain file references based on header keywords.
    Excludes columns that are clearly FILE TYPE / FORMAT columns (e.g. 'File type', 'Format').
    Returns a list of column indices (0-based).
    """
    file_col_indices = []
    for idx, cell in enumerate(header_row):
        cell_norm = _normalize(str(cell) if cell is not None else "")
        has_file_kw = any(kw in cell_norm for kw in _FILE_COL_KEYWORDS)
        # Word-boundary matched: a bare "ext" in cell_norm check wrongly excluded
        # a genuine evidence column titled "Extra Attachments" (contains "ext")
        # as if it were a File Type/Format column.
        has_type_kw = any(_kw_match(kw, cell_norm) for kw in _FILE_TYPE_EXCLUSION_KEYWORDS)
        # Include only if it has a file keyword AND does NOT have a type/format keyword
        if has_file_kw and not has_type_kw:
            file_col_indices.append(idx)
    return file_col_indices


def _get_file_column_roles(header_row: list, file_cols: List[int]) -> Dict[int, str]:
    """
    Classifies each detected file column as 'policy', 'evidence', or 'generic' based
    on its header text (e.g. 'Policy (source-grounded)' -> 'policy', 'File name' ->
    'generic'). Lets downstream code tell the LLM which locked file(s) the auditor
    intended as policy proof vs operational evidence proof, instead of locking both
    into one undifferentiated blob and making the LLM re-derive the split blind.
    """
    roles = {}
    for fc in file_cols:
        col_name_norm = _normalize(str(header_row[fc]) if fc < len(header_row) else "")
        is_policy = "policy" in col_name_norm or "standard" in col_name_norm
        is_evidence = any(kw in col_name_norm for kw in ("evidence", "log", "screenshot", "proof", "report", "result"))
        if is_policy and not is_evidence:
            roles[fc] = "policy"
        elif is_evidence and not is_policy:
            roles[fc] = "evidence"
        else:
            roles[fc] = "generic"
    return roles


_ROW_NUMBER_HEADER_LABELS = {"s no", "sno", "sl no", "sl", "no", "sr no", "srno", "#", "row", "row no", "index", ""}


def _detect_id_column(header_row: list) -> Optional[int]:
    """Detects Control ID column in header row."""
    id_keywords = {"control id", "control #", "iso control", "clause", "control_id", "id", "ref"}
    for idx, cell in enumerate(header_row):
        cell_norm = _normalize(str(cell) if cell is not None else "")
        # Word-boundary matched: a bare "id" in cell_norm check misdetected
        # headers like "Evidence Provided" or "Void" as the Control ID column.
        if any(kw == cell_norm or _kw_match(kw, cell_norm) for kw in id_keywords):
            return idx
    return None


def _detect_name_column(header_row: list) -> Optional[int]:
    """Detects Control Name column in header row."""
    name_keywords = {"control name", "control title", "title", "name", "label"}
    for idx, cell in enumerate(header_row):
        cell_norm = _normalize(str(cell) if cell is not None else "")
        # Exclude columns containing file keywords (e.g. "File name", "Document name")
        if any(file_kw in cell_norm for file_kw in _FILE_COL_KEYWORDS):
            continue
        if any(kw == cell_norm or kw in cell_norm for kw in name_keywords):
            return idx
    return None


def _detect_question_column(header_row: list) -> int:
    """
    Auto-detect which column holds the audit check question/control name.
    Falls back to the first non-file-reference column after the ID column,
    or the ID column itself as an absolute last resort.

    Used to unconditionally fall back to column index 1 -- on a real customer
    checklist with headers "Control ID (ISO)" / "Policy Document Name" /
    "Evidence Document Name", nothing matches a question keyword, so that
    blind fallback claimed column 1 -- the genuine Policy file column --
    as the "question" column. Downstream, parse_excel_scoping_checklist
    excludes whatever this returns from file_cols, so every row's policy
    file silently vanished before it ever reached the matching step.
    """
    question_keywords = {
        "audit check", "question", "check", "requirement", "description",
        "audit question", "check question", "objective", "observation"
    }
    for idx, cell in enumerate(header_row):
        cell_norm = _normalize(str(cell) if cell is not None else "")
        if idx == 0 and cell_norm in _ROW_NUMBER_HEADER_LABELS:
            continue
        for kw in question_keywords:
            if kw in cell_norm:
                return idx

    # No header matched a question keyword -- fall back to the first column
    # (skipping the ID column) that isn't itself a recognized file-reference
    # column, so a genuine Policy/Evidence column is never silently stolen.
    for idx, cell in enumerate(header_row):
        if idx == 0:
            continue
        cell_norm = _normalize(str(cell) if cell is not None else "")
        if any(file_kw in cell_norm for file_kw in _FILE_COL_KEYWORDS):
            continue
        return idx

    # Every other column looked like a file column -- falling back to the ID
    # column is always safe: it's already excluded from file_cols via id_col
    # separately, and the caller skips using it as question text when it
    # collides with id_col (see the question_col != id_col check below).
    return 0


def parse_excel_scoping_checklist(
    file_path: str,
    sheet_name: str = None,
    uploaded_filenames: List[str] = None,
    framework: str = None
) -> List[Dict]:
    """
    Parse an Excel scoping checklist and return a list of checklist items.
    Supports all 3 input styles: ID only, Name only, Question only, and combinations.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel scoping file not found: {file_path}")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        preferred_names = ["Audit Check", "Scoping", "Checklist", "Controls",
                           "Sheet1", "Sheet 1", "Data"]
        target_sheet = None

        if sheet_name and sheet_name in wb.sheetnames:
            target_sheet = wb[sheet_name]
        else:
            for name in preferred_names:
                if name in wb.sheetnames:
                    target_sheet = wb[name]
                    break
            if target_sheet is None:
                target_sheet = wb[wb.sheetnames[0]]

        rows = list(target_sheet.iter_rows(values_only=True))
        if not rows:
            return []

        header_row_idx = 0
        for i, row in enumerate(rows):
            if any(cell is not None and str(cell).strip() for cell in row):
                header_row_idx = i
                break

        header_row = [str(c).strip() if c is not None else "" for c in rows[header_row_idx]]

        id_col = _detect_id_column(header_row)
        name_col = _detect_name_column(header_row)
        question_col = _detect_question_column(header_row)
        file_cols = _detect_file_columns(header_row)

        file_cols = [c for c in file_cols if c not in (id_col, name_col, question_col)]
        file_col_roles = _get_file_column_roles(header_row, file_cols)

        use_cases = _load_use_cases()
        # Confine matching to the framework this audit is running, so a row with
        # no control ID cannot resolve into a different standard's control set.
        if framework:
            _before = len(use_cases)
            use_cases = filter_use_cases_by_framework(use_cases, framework)
            if len(use_cases) != _before:
                print(f"[EXCEL PARSER] Framework '{framework}': matching against "
                      f"{len(use_cases)} of {_before} controls.", flush=True)
        items = []

        for row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
            if not any(c is not None and str(c).strip() for c in row):
                continue

            id_text = str(row[id_col]).strip() if id_col is not None and id_col < len(row) and row[id_col] is not None else ""
            name_text = str(row[name_col]).strip() if name_col is not None and name_col < len(row) and row[name_col] is not None else ""
            # question_col == id_col only when _detect_question_column() hit its
            # absolute-last-resort fallback (every other column was a file
            # column) -- treat that as "no question column" rather than
            # duplicating the raw control ID as a fake requirement question,
            # so the framework's own expected_evidence/control_label is used
            # instead (see req_q resolution below).
            q_text = str(row[question_col]).strip() if question_col != id_col and question_col < len(row) and row[question_col] is not None else ""

            resolution_text = " ".join(t for t in [id_text, name_text, q_text] if t)
            if not resolution_text or _normalize(resolution_text) in {
                "audit check", "question", "control", "sl no", "s.no", "sno",
                "control control name", "control name"
            }:
                continue

            raw_files, raw_policy_files, raw_evidence_files = [], [], []
            for fc in file_cols:
                if fc < len(row) and row[fc] is not None:
                    val = str(row[fc]).strip()
                    if val and val.lower() not in {"file name", "file", "filename", "n/a", "na", "-", ""}:
                        raw_files.append(val)
                        role = file_col_roles.get(fc, "generic")
                        if role == "policy":
                            raw_policy_files.append(val)
                        elif role == "evidence":
                            raw_evidence_files.append(val)

            if uploaded_filenames:
                matched_files = _match_filenames(raw_files, uploaded_filenames)
                matched_policy_files = _match_filenames(raw_policy_files, uploaded_filenames)
                matched_evidence_files = _match_filenames(raw_evidence_files, uploaded_filenames)
            else:
                matched_files = raw_files
                matched_policy_files = raw_policy_files
                matched_evidence_files = raw_evidence_files

            ctrl_info = _resolve_control(id_text, name_text, q_text, use_cases)

            # Determine requirement question & metadata
            if q_text:
                req_q = q_text
                req_source = "excel"
                req_status = "RESOLVED"
            elif ctrl_info["resolved"] and ctrl_info.get("expected_evidence"):
                req_q = ctrl_info["expected_evidence"]
                req_source = "framework"
                req_status = "RESOLVED"
            elif ctrl_info["resolved"]:
                req_q = ctrl_info["control_label"]
                req_source = "framework"
                req_status = "RESOLVED"
            else:
                req_q = None
                req_source = "unresolved"
                req_status = "UNRESOLVED"

            items.append({
                "row_index":                    row_idx,
                "question":                     req_q or ctrl_info["control_label"],
                "requirement_question":        req_q,
                "requirement_question_source": req_source,
                "requirement_question_status": req_status,
                "files":                        matched_files,
                "policy_files":                 matched_policy_files,
                "evidence_files":               matched_evidence_files,
                "raw_file_refs":                raw_files,
                "control_id":                   ctrl_info["control_id"],
                "control_name":                 ctrl_info["control_name"],
                "control_label":                ctrl_info["control_label"],
                "expected_evidence":            ctrl_info["expected_evidence"],
                "prompt_hint":                  ctrl_info["prompt_hint"],
                "severity":                     ctrl_info["severity"],
            })

        print(f"[EXCEL PARSER] Parsed {len(items)} checklist items.", flush=True)
        return items
    finally:
        try:
            wb.close()
        except Exception:
            pass



_FILENAME_EXT_RE = re.compile(r'\.(docx?|pdf|xlsx?|csv|pptx?|txt|png|jpe?g|zip)\b', re.IGNORECASE)
_CITATION_SPLIT_RE = re.compile(r'\s*;\s*|\n')
_CITATION_LEADING_TAG_RE = re.compile(r'^\[[^\]]+\]\s*')
# Marks a segment as citation-shaped rather than a narrative sentence fragment:
# a version number ("V17.0", "v6.3"), a page/slide reference ("p.2", "pp.5-6",
# "slides 5, 9"), or a "... section" suffix -- the actual short-citation
# patterns real auditors use in these columns (confirmed against a real
# customer checklist), as opposed to a clause split out of a longer sentence
# describing what a policy requires (e.g. "badges are required.").
_CITATION_MARKER_RE = re.compile(
    r'\bv\d+(\.\d+)+\b|\bp{1,2}\.\s*\d|\bslides?\s+\d|\bsection\b',
    re.IGNORECASE
)


def _looks_like_filename_reference(segment: str) -> bool:
    """
    A segment is only treated as a candidate file citation (matched against
    uploads, or kept as an unresolved reference) if it's filename-shaped -- has
    a recognized extension, or is short AND carries some other citation marker
    (version number, page/slide reference, "... section") like "Some Policy
    V17.0 -- p.2." Short prose clauses split out of a longer sentence by
    _split_citation_segments (e.g. "badges are required.", cut from "...
    visitors must be escorted; badges are required.") also land under the
    length cap but aren't citations at all -- confirmed against a real
    customer checklist where narrative Policy-column text produced fake
    "missing file" placeholders like that instead of being correctly
    discarded. Long narrative text with no extension is never kept as a fake
    "locked filename" -- that silently breaks retrieval later, since no real
    file will ever match it.
    """
    if _FILENAME_EXT_RE.search(segment):
        return True
    return len(segment) <= 80 and bool(_CITATION_MARKER_RE.search(segment))


def _split_citation_segments(raw_ref: str) -> List[str]:
    """
    A single cell may cite multiple files at once (e.g. "A.docx; B.pdf -- p.2."),
    or bracket-tag a citation (e.g. "[Published] Some Policy.docx"). Splits on
    ';' / newlines and strips leading bracket tags so each file gets matched
    independently instead of the whole cell being treated as one reference.
    """
    parts = [p.strip() for p in _CITATION_SPLIT_RE.split(raw_ref) if p.strip()]
    cleaned = [_CITATION_LEADING_TAG_RE.sub('', p).strip() for p in parts]
    cleaned = [p for p in cleaned if p]
    return cleaned or [raw_ref]


def _match_filenames(
    raw_refs: List[str],
    uploaded_filenames: List[str]
) -> List[str]:
    """
    Fuzzy-match raw Excel file references to actual uploaded filenames.

    Each raw reference may cite multiple files in one cell, or be pure
    narrative text with no file reference at all -- both are handled by first
    splitting into citation segments (see _split_citation_segments), then
    matching each segment independently.

    Strategy per segment:
    1. Exact match (case-insensitive, extension-stripped)
    2. Partial match -- checks BOTH the full uploaded filename and its
       extension-stripped stem as a substring, since citation text doesn't
       always repeat the file extension (e.g. "...Policy V17.0 -- p.2." cites
       "...Policy V17.0.pdf" without ever writing ".pdf")
    3. If nothing matches, keep the segment only if it's filename-shaped (see
       _looks_like_filename_reference) so the auditor can see an expected file
       that hasn't been uploaded yet. Pure narrative segments are dropped
       instead of being kept as a fake filename reference.
    """
    def _resolve_segment(seg: str) -> Optional[str]:
        seg_norm = _normalize(seg)

        # Step 1: exact match (extension-stripped)
        for upl in uploaded_filenames:
            upl_stem_norm = _normalize(os.path.splitext(upl)[0])
            if seg_norm == upl_stem_norm or seg_norm == _normalize(upl):
                return upl

        # Step 2: partial match -- try both the full filename and its
        # extension-stripped stem, since citations often omit the extension
        best_match, best_len = None, 0
        for upl in uploaded_filenames:
            for candidate in (_normalize(upl), _normalize(os.path.splitext(upl)[0])):
                if not candidate:
                    continue
                if candidate in seg_norm and len(candidate) > best_len:
                    best_len, best_match = len(candidate), upl
                elif seg_norm in candidate and len(seg_norm) > best_len:
                    best_len, best_match = len(seg_norm), upl
        return best_match

    matched = []
    for ref in raw_refs:
        for seg in _split_citation_segments(ref):
            resolved = _resolve_segment(seg) if uploaded_filenames else None
            if resolved:
                matched.append(resolved)
            elif _looks_like_filename_reference(seg):
                # Step 3: keep as an unresolved-but-plausible reference
                matched.append(seg)

    # Deduplicate while preserving order -- the same file can legitimately be
    # cited twice (e.g. once in the Policy column, once in Evidence)
    seen = set()
    result = []
    for f in matched:
        if f not in seen:
            seen.add(f)
            result.append(f)
    return result


def get_locked_filenames_for_control(
    checklist_items: List[Dict],
    control_id: str
) -> List[str]:
    """
    Returns all locked filenames for a given control_id across all checklist items.
    Useful when multiple checklist rows map to the same control (e.g., two NTP rows → 8.17).
    """
    files = []
    for item in checklist_items:
        if item.get("control_id") == control_id:
            files.extend(item.get("files", []))
    # Deduplicate while preserving order
    seen = set()
    result = []
    for f in files:
        if f not in seen:
            seen.add(f)
            result.append(f)
    return result
